"""Tests for the export-detalle-mes-xlsx feature.

Covers: happy path, empty month, missing mono (404), unauthenticated (302),
numeric cell types, credit-note (NC) negative importe.
"""
from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

import pytest

from website.models import Categoria, Factura, Monotributista, db


# ---------------------------------------------------------------------------
# Helpers shared across tests
# ---------------------------------------------------------------------------

def login(client):
    return client.post(
        "/login",
        data={"username": "admin", "password": "admin"},
        follow_redirects=True,
    )


def create_categoria(app):
    with app.app_context():
        categoria = Categoria(nombre="A", orden=1, tope_facturacion=Decimal("0.00"))
        db.session.add(categoria)
        db.session.commit()
        return categoria.id


def create_mono(app, categoria_id, *, razon_social="Test SRL", cuit="20123456789"):
    with app.app_context():
        mono = Monotributista(
            razon_social=razon_social,
            cuit=cuit,
            clave_fiscal="clave",
            categoria_actual_id=categoria_id,
            categoria_corresponde_id=categoria_id,
        )
        db.session.add(mono)
        db.session.commit()
        return mono.id


def create_factura(app, mono_id, *, fecha, importe, tipo_comp="B",
                   fecha_desde=None, fecha_hasta=None, numero_comp="00001-00000001"):
    with app.app_context():
        f = Factura(
            monotributista_id=mono_id,
            fecha=fecha,
            tipo_comp=tipo_comp,
            numero_comp=numero_comp,
            importe_total=importe,
            fecha_desde=fecha_desde or fecha,
            fecha_hasta=fecha_hasta or fecha,
        )
        db.session.add(f)
        db.session.commit()


# ---------------------------------------------------------------------------
# T2a — Happy path: 200, correct content-type, parseable workbook, structure
# ---------------------------------------------------------------------------

def test_export_xlsx_happy_path(client, app):
    """Happy path: returns 200, xlsx content-type, parseable workbook."""
    import openpyxl

    cat_id = create_categoria(app)
    mono_id = create_mono(app, cat_id)
    # Factura in Jan 2025
    create_factura(app, mono_id, fecha=date(2025, 1, 15), importe=Decimal("5000.00"),
                   fecha_desde=date(2025, 1, 15), fecha_hasta=date(2025, 1, 15))

    login(client)
    resp = client.get(f"/api/calculo/{mono_id}/mes/2025/1/export.xlsx?anchor=2025-01")
    assert resp.status_code == 200
    assert resp.content_type.startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws = wb.active

    # Row 1 is the header row
    headers = [ws.cell(1, col).value for col in range(1, 7)]
    assert headers == [
        "Nro Comprobante", "Tipo", "Desde", "Hasta", "Importe Total", "Importe Mes"
    ]

    # At least one data row (row 2)
    assert ws.max_row >= 3  # header + data + total

    # Last row has "Total" in col E (column 5)
    last_row = ws.max_row
    assert ws.cell(last_row, 5).value == "Total"

    # Total Importe Mes cell is numeric (float/int), not str
    total_value = ws.cell(last_row, 6).value
    assert isinstance(total_value, (int, float)), f"Expected numeric, got {type(total_value)}"


# ---------------------------------------------------------------------------
# T2b — Empty month: 200, valid xlsx, note in A2, Total = 0.0
# ---------------------------------------------------------------------------

def test_export_xlsx_empty_month(client, app):
    """Empty month: 200, valid xlsx, note row, Total Importe Mes = 0.0."""
    import openpyxl

    cat_id = create_categoria(app)
    mono_id = create_mono(app, cat_id)
    # No invoices at all — any month is empty

    login(client)
    resp = client.get(f"/api/calculo/{mono_id}/mes/2025/6/export.xlsx?anchor=2025-06")
    assert resp.status_code == 200
    assert resp.content_type.startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws = wb.active

    # Header row 1
    assert ws.cell(1, 1).value == "Nro Comprobante"

    # Note in A2
    assert ws.cell(2, 1).value == "Sin comprobantes para este mes"

    # Total row: col F = 0.0
    last_row = ws.max_row
    assert ws.cell(last_row, 5).value == "Total"
    assert ws.cell(last_row, 6).value == 0.0


# ---------------------------------------------------------------------------
# T2c — Missing monotributista → 404
# ---------------------------------------------------------------------------

def test_export_xlsx_missing_mono_returns_404(client, app):
    """Non-existent mono_id must return 404."""
    login(client)
    resp = client.get("/api/calculo/99999/mes/2025/1/export.xlsx")
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# T2d — Unauthenticated → redirect (302)
# ---------------------------------------------------------------------------

def test_export_xlsx_unauthenticated_redirects(client, app):
    """Unauthenticated request must redirect to login (302)."""
    resp = client.get("/api/calculo/1/mes/2025/1/export.xlsx")
    assert resp.status_code == 302


# ---------------------------------------------------------------------------
# T2e — Numeric cell types: importe_total and importe_mes must be float/int
# ---------------------------------------------------------------------------

def test_export_xlsx_numeric_cell_types(client, app):
    """Importe columns must be numeric cells, not strings."""
    import openpyxl

    cat_id = create_categoria(app)
    mono_id = create_mono(app, cat_id)
    create_factura(app, mono_id, fecha=date(2025, 3, 10), importe=Decimal("1234.56"),
                   fecha_desde=date(2025, 3, 10), fecha_hasta=date(2025, 3, 10))

    login(client)
    resp = client.get(f"/api/calculo/{mono_id}/mes/2025/3/export.xlsx?anchor=2025-03")
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws = wb.active

    # Iterate data rows (row 2 up to but not including total row)
    for row_idx in range(2, ws.max_row):
        col_e = ws.cell(row_idx, 5).value
        col_f = ws.cell(row_idx, 6).value
        assert isinstance(col_e, (int, float)), (
            f"Row {row_idx} col E expected numeric, got {type(col_e)}: {col_e!r}"
        )
        assert isinstance(col_f, (int, float)), (
            f"Row {row_idx} col F expected numeric, got {type(col_f)}: {col_f!r}"
        )


# ---------------------------------------------------------------------------
# T2f — NC row has negative importe (not double-negated)
# ---------------------------------------------------------------------------

def test_export_xlsx_nc_row_has_negative_importe(client, app):
    """Credit note row must have negative importe in xlsx (no double-negation)."""
    import openpyxl

    cat_id = create_categoria(app)
    mono_id = create_mono(app, cat_id)
    # Create a credit note — the app stores it as negative automatically when tipo_comp starts with NC
    # We pass a positive importe here; views.py create_factura route negates it.
    # But for the DB fixture we store it already negative as the app would store it.
    with app.app_context():
        f = Factura(
            monotributista_id=mono_id,
            fecha=date(2025, 4, 5),
            tipo_comp="NCB",
            numero_comp="00001-00000010",
            importe_total=Decimal("-500.00"),  # already negative as stored by app
            fecha_desde=date(2025, 4, 5),
            fecha_hasta=date(2025, 4, 5),
        )
        db.session.add(f)
        db.session.commit()

    login(client)
    resp = client.get(f"/api/calculo/{mono_id}/mes/2025/4/export.xlsx?anchor=2025-04")
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws = wb.active

    # Find the NCB row (row 2 if only one invoice)
    found_negative = False
    for row_idx in range(2, ws.max_row):
        tipo = ws.cell(row_idx, 2).value
        if tipo and str(tipo).upper().startswith("NC"):
            col_f = ws.cell(row_idx, 6).value
            assert isinstance(col_f, (int, float)), f"Expected numeric, got {type(col_f)}"
            assert col_f < 0, f"Expected negative importe for NC row, got {col_f}"
            found_negative = True

    assert found_negative, "No NC row found in the xlsx"
